#!/usr/bin/env python
# -*- coding: utf8 -*-
import openpyxl

#新規作成
wb = openpyxl.Workbook()

sheet = wb.active
sheet.title

id = 'TCG Tournament TEST'

sheet.title = id
test1 = ["Test Upload,"Test Upload","3","4"]
test2 = [" for pull request","p2","p3","p4"]

i = 1
for t1 in test1:
    sheet.cell(row=i,column=2).value = t1
    i = i+1

k = 1
for t2 in test2:
    sheet.cell(row=i,column=1).value = t2
    k = k+1

#保存
wb.save(id + '.xls')